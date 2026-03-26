"""
iBizSim 登录验证 GUI 工具
提供图形界面用于输入账号和密码，无需配置环境变量
"""
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.config.settings import Settings
from src.browser.browser_manager import BrowserManager
from src.browser.page_handler import PageHandler
from src.auth.login_handler import LoginHandler
from src.data.game_extractor import GameExtractor
from src.utils.logger import get_logger


class LoginGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("iBizSim 助手")
        self.root.state('zoomed')
        self.root.resizable(True, True)
        
        self.username = tk.StringVar()
        self.password = tk.StringVar()
        self.show_password = tk.BooleanVar(value=True)
        self.status = tk.StringVar(value="准备就绪")
        self.is_running = False
        self.team_name = ""
        self.excel_file_path = ""
        
        self.games = []
        self.selected_game = None
        self.navigation_queue = []
        self.playwright_thread = None
        self.playwright_queue = []
        self.playwright_running = False
        
        self.setup_ui()
        self.logger = get_logger()
        
        self.browser_manager = None
        self.page_handler = None
        self.login_handler = None
        self.game_extractor = None
        self.settings = None
        self.current_game_id = None
        self.current_team_id = None
        self.period_8_url = None
    
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        title_label = ttk.Label(
            main_frame, 
            text="iBizSim 助手",
            font=("Microsoft YaHei UI", 16, "bold"),
            foreground="#2c3e50"
        )
        title_label.pack(pady=(0, 5))
        
        columns_frame = ttk.Frame(main_frame)
        columns_frame.pack(fill=tk.BOTH, expand=True)
        
        columns_frame.columnconfigure(0, weight=1)
        columns_frame.columnconfigure(1, weight=1)
        columns_frame.columnconfigure(2, weight=1)
        
        column1 = ttk.Frame(columns_frame)
        column1.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        
        column2 = ttk.Frame(columns_frame)
        column2.grid(row=0, column=1, sticky="nsew", padx=5)
        
        column3 = ttk.Frame(columns_frame)
        column3.grid(row=0, column=2, sticky="nsew", padx=(5, 0))
        
        input_frame = ttk.LabelFrame(column1, text="登录信息", padding="10")
        input_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(input_frame, text="用户名:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        username_entry = ttk.Entry(
            input_frame, 
            textvariable=self.username,
            width=25,
            font=("Microsoft YaHei UI", 10)
        )
        username_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(input_frame, text="密码:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        password_entry = ttk.Entry(
            input_frame,
            textvariable=self.password,
            width=18,
            font=("Microsoft YaHei UI", 10),
            show=""
        )
        password_entry.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        
        self.password_entry = password_entry
        
        show_password_check = ttk.Checkbutton(
            input_frame,
            text="显示密码",
            variable=self.show_password,
            command=self.toggle_password_visibility
        )
        show_password_check.grid(row=1, column=2, sticky=tk.W, padx=(5, 0), pady=5)
        
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=2, column=0, columnspan=3, pady=(10, 0))
        
        login_button = ttk.Button(
            button_frame,
            text="登录",
            command=self.start_verification,
            width=12,
            state=tk.NORMAL
        )
        login_button.pack(side=tk.LEFT, padx=2)
        
        stop_button = ttk.Button(
            button_frame,
            text="停止",
            command=self.stop_verification,
            width=12,
            state=tk.DISABLED
        )
        stop_button.pack(side=tk.LEFT, padx=2)
        
        clear_button = ttk.Button(
            button_frame,
            text="清空",
            command=self.clear_inputs,
            width=12
        )
        clear_button.pack(side=tk.LEFT, padx=2)
        
        status_frame = ttk.LabelFrame(column1, text="状态", padding="10")
        status_frame.pack(fill=tk.X, pady=(10, 0))
        
        status_entry = ttk.Entry(
            status_frame,
            textvariable=self.status,
            font=("Microsoft YaHei UI", 10),
            state='readonly'
        )
        status_entry.pack(fill=tk.X)
        
        games_frame = ttk.LabelFrame(column1, text="我参加的赛事", padding="10")
        games_frame.pack(fill=tk.BOTH, expand=False, pady=(10, 0))
        
        games_listbox_frame = ttk.Frame(games_frame)
        games_listbox_frame.pack(fill=tk.BOTH, expand=True)
        
        self.games_listbox = tk.Listbox(
            games_listbox_frame,
            height=21,
            font=("Microsoft YaHei UI", 10),
            selectmode=tk.SINGLE
        )
        self.games_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        v_scrollbar = ttk.Scrollbar(games_listbox_frame, orient=tk.VERTICAL, command=self.games_listbox.yview)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.games_listbox.config(yscrollcommand=v_scrollbar.set)
        
        h_scrollbar = ttk.Scrollbar(games_frame, orient=tk.HORIZONTAL, command=self.games_listbox.xview)
        h_scrollbar.pack(fill=tk.X, pady=(5, 0))
        self.games_listbox.config(xscrollcommand=h_scrollbar.set)
        
        enter_game_button = ttk.Button(
            games_frame,
            text="进入比赛",
            command=self.enter_game,
            state=tk.DISABLED
        )
        enter_game_button.pack(fill=tk.X, pady=(5, 0))
        
        log_button_frame = ttk.Frame(column1)
        log_button_frame.pack(fill=tk.X, pady=(10, 0))
        
        log_button = ttk.Button(
            log_button_frame,
            text="查看日志",
            command=self.show_log_dialog,
            width=15
        )
        log_button.pack(side=tk.LEFT)
        
        exit_button = ttk.Button(
            log_button_frame,
            text="退出",
            command=self.on_closing,
            width=15
        )
        exit_button.pack(side=tk.RIGHT)
        
        file_frame = ttk.LabelFrame(column2, text="文件选择", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        file_row_frame = ttk.Frame(file_frame)
        file_row_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(file_row_frame, text="文件地址:").pack(side=tk.LEFT, padx=(0, 5))
        
        file_path_var = tk.StringVar()
        file_entry = ttk.Entry(
            file_row_frame,
            textvariable=file_path_var,
            font=("Microsoft YaHei UI", 10)
        )
        file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        open_button = ttk.Button(
            file_row_frame,
            text="打开",
            command=lambda: self.open_file(file_path_var),
            width=10
        )
        open_button.pack(side=tk.LEFT)
        
        self.file_path_var = file_path_var
        
        notebook = ttk.Notebook(column2)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        self.notebook = notebook
        
        rules_tab = ttk.Frame(notebook)
        notebook.add(rules_tab, text="规则详情")
        
        rules_button_frame = ttk.Frame(rules_tab)
        rules_button_frame.pack(fill=tk.X, pady=(5, 5))
        
        copy_button = ttk.Button(
            rules_button_frame,
            text="复制规则",
            command=self.copy_rules,
            state=tk.DISABLED,
            width=10
        )
        copy_button.pack(side=tk.LEFT, padx=(0, 2))
        
        import_button = ttk.Button(
            rules_button_frame,
            text="导入规则",
            command=self.import_rules,
            state=tk.DISABLED,
            width=10
        )
        import_button.pack(side=tk.LEFT, padx=2)
        
        rules_split_frame = ttk.Frame(rules_tab)
        rules_split_frame.pack(fill=tk.BOTH, expand=True)
        
        rules_split_frame.rowconfigure(0, weight=1)
        rules_split_frame.columnconfigure(0, weight=3)
        rules_split_frame.columnconfigure(1, weight=1)
        
        rules_left_frame = ttk.Frame(rules_split_frame)
        rules_left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 1))
        
        rules_right_frame = ttk.Frame(rules_split_frame)
        rules_right_frame.grid(row=0, column=1, sticky="nsew", padx=(1, 0))
        
        rules_right_frame.rowconfigure(0, weight=1)
        rules_right_frame.columnconfigure(0, weight=1)
        
        rules_right_canvas = tk.Canvas(rules_right_frame)
        rules_right_scrollbar = ttk.Scrollbar(rules_right_frame, orient="vertical", command=rules_right_canvas.yview)
        rules_right_scrollable_frame = ttk.Frame(rules_right_canvas)
        
        rules_right_scrollable_frame.bind(
            "<Configure>",
            lambda e: rules_right_canvas.configure(scrollregion=rules_right_canvas.bbox("all"))
        )
        
        def on_canvas_configure(event):
            canvas_width = event.width
            canvas_height = event.height
            rules_right_canvas.itemconfig(1, width=canvas_width)
            rules_right_canvas.itemconfig(1, height=canvas_height)
        
        rules_right_canvas.bind("<Configure>", on_canvas_configure)
        
        rules_right_canvas.create_window((0, 0), window=rules_right_scrollable_frame, anchor="nw")
        rules_right_canvas.configure(yscrollcommand=rules_right_scrollbar.set)
        
        rules_right_canvas.grid(row=0, column=0, sticky="nsew")
        rules_right_scrollbar.grid(row=0, column=1, sticky="ns")
        
        rules_table_frame = ttk.Frame(rules_left_frame)
        rules_table_frame.pack(fill=tk.BOTH, expand=False)
        
        rules_table_frame.columnconfigure(0, weight=1)
        rules_table_frame.rowconfigure(0, weight=1)
        
        rules_table = ttk.Treeview(
            rules_table_frame,
            columns=("col1", "col2"),
            show="headings",
            height=25
        )
        rules_table.heading("col1", text="参数")
        rules_table.heading("col2", text="值")
        rules_table.column("col1", width=50, anchor=tk.CENTER)
        rules_table.column("col2", width=80, anchor=tk.W)
        rules_table.grid(row=0, column=0, sticky="nsew")
        
        table_scrollbar = ttk.Scrollbar(rules_table_frame, orient=tk.VERTICAL, command=rules_table.yview)
        table_scrollbar.grid(row=0, column=1, sticky="ns")
        rules_table.config(yscrollcommand=table_scrollbar.set)
        
        # 添加固定参数行
        parameters = [
            "当期可运输比例",
            "公司总数",
            "公司序号",
            "原材料库存费用",
            "购机费用",
            "原材料固定运费",
            "原材料变动运费",
            "原材料可用比例",
            "维修费",
            "新员工培训费",
            "安置费",
            "基本工资",
            "一加特殊工资",
            "二班正班工资",
            "二加特殊工资",
            "废品系数",
            "最高工资系数",
            "最低资金额度",
            "贷款利息",
            "国债利息",
            "债券利息",
            "税收比例",
            "减税比例",
            "资金有效性",
            "本期利润",
            "市场份额",
            "累计分红",
            "累计缴税",
            "净资产",
            "人均利润率",
            "资本利润率"
        ]
        
        for param in parameters:
            rules_table.insert("", tk.END, values=(param, ""))
        
        self.rules_table = rules_table
        
        discount_frame = ttk.Frame(rules_left_frame)
        discount_frame.pack(fill=tk.BOTH, expand=False, pady=(10, 0))
        
        discount_frame.columnconfigure(0, weight=1)
        discount_frame.rowconfigure(0, weight=1)
        
        discount_table = ttk.Treeview(
            discount_frame,
            columns=("col1", "col2"),
            show="headings",
            height=3
        )
        discount_table.heading("col1", text="")
        discount_table.heading("col2", text="原材料折扣")
        discount_table.column("col1", width=90, anchor=tk.CENTER)
        discount_table.column("col2", width=40, anchor=tk.W)
        discount_table.grid(row=0, column=0, sticky="nsew")
        
        discount_scrollbar = ttk.Scrollbar(discount_frame, orient=tk.VERTICAL, command=discount_table.yview)
        discount_scrollbar.grid(row=0, column=1, sticky="ns")
        discount_table.config(yscrollcommand=discount_scrollbar.set)
        
        discount_items = ["", "", ""]
        for item in discount_items:
            discount_table.insert("", tk.END, values=(item, ""))
        
        self.discount_table = discount_table
        
        product_table_frame = ttk.Frame(rules_right_scrollable_frame)
        product_table_frame.pack(fill=tk.X, expand=False)
        
        product_table = ttk.Treeview(
            product_table_frame,
            columns=("col1", "col2", "col3", "col4", "col5"),
            show="headings",
            height=4
        )
        product_table.heading("col1", text="")
        product_table.heading("col2", text="成品库存费")
        product_table.heading("col3", text="")
        product_table.heading("col4", text="一正管理费")
        product_table.heading("col5", text="二正管理费")
        product_table.column("col1", width=80, anchor=tk.CENTER, stretch=False)
        product_table.column("col2", width=100, anchor=tk.CENTER, stretch=True)
        product_table.column("col3", width=80, anchor=tk.CENTER, stretch=False)
        product_table.column("col4", width=100, anchor=tk.CENTER, stretch=True)
        product_table.column("col5", width=100, anchor=tk.CENTER, stretch=True)
        product_table.pack(fill=tk.X, expand=False)
        
        products = ["产品1", "产品2", "产品3", "产品4"]
        for product in products:
            product_table.insert("", tk.END, values=(product, "", product, "", ""))
        
        self.product_table = product_table
        
        conversion_table_frame = ttk.Frame(rules_right_scrollable_frame)
        conversion_table_frame.pack(fill=tk.X, expand=False)
        
        conversion_table = ttk.Treeview(
            conversion_table_frame,
            columns=("col1", "col2", "col3", "col4", "col5"),
            show="headings",
            height=4
        )
        conversion_table.heading("col1", text="订货转化比例")
        conversion_table.heading("col2", text="产品1")
        conversion_table.heading("col3", text="产品2")
        conversion_table.heading("col4", text="产品3")
        conversion_table.heading("col5", text="产品4")
        conversion_table.column("col1", width=100, anchor=tk.CENTER, stretch=False)
        conversion_table.column("col2", width=80, anchor=tk.CENTER, stretch=True)
        conversion_table.column("col3", width=80, anchor=tk.CENTER, stretch=True)
        conversion_table.column("col4", width=80, anchor=tk.CENTER, stretch=True)
        conversion_table.column("col5", width=80, anchor=tk.CENTER, stretch=True)
        conversion_table.pack(fill=tk.X, expand=False)
        
        markets = ["市场1", "市场2", "市场3", "市场4"]
        for market in markets:
            conversion_table.insert("", tk.END, values=(market, "", "", "", ""))
        
        self.conversion_table = conversion_table
        
        production_table_frame = ttk.Frame(rules_right_scrollable_frame)
        production_table_frame.pack(fill=tk.X, expand=False)
        
        production_table = ttk.Treeview(
            production_table_frame,
            columns=("col1", "col2", "col3", "col4", "col5"),
            show="headings",
            height=3
        )
        production_table.heading("col1", text="产品生产消耗")
        production_table.heading("col2", text="产品1")
        production_table.heading("col3", text="产品2")
        production_table.heading("col4", text="产品3")
        production_table.heading("col5", text="产品4")
        production_table.column("col1", width=120, anchor=tk.CENTER, stretch=False)
        production_table.column("col2", width=80, anchor=tk.CENTER, stretch=True)
        production_table.column("col3", width=80, anchor=tk.CENTER, stretch=True)
        production_table.column("col4", width=80, anchor=tk.CENTER, stretch=True)
        production_table.column("col5", width=80, anchor=tk.CENTER, stretch=True)
        production_table.pack(fill=tk.X, expand=False)
        
        production_items = ["机器（时）", "人力（时）", "原材料"]
        for item in production_items:
            production_table.insert("", tk.END, values=(item, "", "", "", ""))
        
        self.production_table = production_table
        
        grade_table_frame = ttk.Frame(rules_right_scrollable_frame)
        grade_table_frame.pack(fill=tk.X, expand=False)
        
        grade_table = ttk.Treeview(
            grade_table_frame,
            columns=("col1", "col2", "col3", "col4", "col5", "col6"),
            show="headings",
            height=4
        )
        grade_table.heading("col1", text="研发投入")
        grade_table.heading("col2", text="等级1")
        grade_table.heading("col3", text="等级2")
        grade_table.heading("col4", text="等级3")
        grade_table.heading("col5", text="等级4")
        grade_table.heading("col6", text="等级5")
        grade_table.column("col1", width=80, anchor=tk.CENTER, stretch=False)
        grade_table.column("col2", width=80, anchor=tk.CENTER, stretch=True)
        grade_table.column("col3", width=80, anchor=tk.CENTER, stretch=True)
        grade_table.column("col4", width=80, anchor=tk.CENTER, stretch=True)
        grade_table.column("col5", width=80, anchor=tk.CENTER, stretch=True)
        grade_table.column("col6", width=80, anchor=tk.CENTER, stretch=True)
        grade_table.pack(fill=tk.X, expand=False)
        
        products = ["产品1", "产品2", "产品3", "产品4"]
        for product in products:
            grade_table.insert("", tk.END, values=(product, "", "", "", "", ""))
        
        self.grade_table = grade_table
        
        shipping_table_frame = ttk.Frame(rules_right_scrollable_frame)
        shipping_table_frame.pack(fill=tk.X, expand=False)
        
        shipping_table = ttk.Treeview(
            shipping_table_frame,
            columns=("col1", "col2", "col3", "col4", "col5"),
            show="headings",
            height=4
        )
        shipping_table.heading("col1", text="固定运费")
        shipping_table.heading("col2", text="市场1")
        shipping_table.heading("col3", text="市场2")
        shipping_table.heading("col4", text="市场3")
        shipping_table.heading("col5", text="市场4")
        shipping_table.column("col1", width=100, anchor=tk.CENTER, stretch=False)
        shipping_table.column("col2", width=80, anchor=tk.CENTER, stretch=True)
        shipping_table.column("col3", width=80, anchor=tk.CENTER, stretch=True)
        shipping_table.column("col4", width=80, anchor=tk.CENTER, stretch=True)
        shipping_table.column("col5", width=80, anchor=tk.CENTER, stretch=True)
        shipping_table.pack(fill=tk.X, expand=False)
        
        shipping_products = ["产品1", "产品2", "产品3", "产品4"]
        for product in shipping_products:
            shipping_table.insert("", tk.END, values=(product, "", "", "", ""))
        
        self.shipping_table = shipping_table
        
        shipping_table3_frame = ttk.Frame(rules_right_scrollable_frame)
        shipping_table3_frame.pack(fill=tk.X, expand=False)
        
        shipping_table3 = ttk.Treeview(
            shipping_table3_frame,
            columns=("col1", "col2", "col3", "col4", "col5"),
            show="headings",
            height=4
        )
        shipping_table3.heading("col1", text="变动运费")
        shipping_table3.heading("col2", text="市场1")
        shipping_table3.heading("col3", text="市场2")
        shipping_table3.heading("col4", text="市场3")
        shipping_table3.heading("col5", text="市场4")
        shipping_table3.column("col1", width=100, anchor=tk.CENTER, stretch=False)
        shipping_table3.column("col2", width=80, anchor=tk.CENTER, stretch=True)
        shipping_table3.column("col3", width=80, anchor=tk.CENTER, stretch=True)
        shipping_table3.column("col4", width=80, anchor=tk.CENTER, stretch=True)
        shipping_table3.column("col5", width=80, anchor=tk.CENTER, stretch=True)
        shipping_table3.pack(fill=tk.X, expand=False)
        
        shipping_products3 = ["产品1", "产品2", "产品3", "产品4"]
        for product in shipping_products3:
            shipping_table3.insert("", tk.END, values=(product, "", "", "", ""))
        
        self.shipping_table3 = shipping_table3
        

        
        salary_coefficient_tab = ttk.Frame(notebook)
        notebook.add(salary_coefficient_tab, text="工资系数")
        self.salary_coefficient_tab_index = notebook.index(salary_coefficient_tab)
        notebook.tab(self.salary_coefficient_tab_index, state="disabled")
        
        salary_button_frame = ttk.Frame(salary_coefficient_tab)
        salary_button_frame.pack(fill=tk.X, pady=(5, 5))
        
        extract_button = ttk.Button(
            salary_button_frame,
            text="提取前八期正品率",
            command=self.extract_quality_rates,
            state=tk.DISABLED,
            width=15
        )
        extract_button.pack(side=tk.LEFT)
        
        initial_data_tab = ttk.Frame(notebook)
        notebook.add(initial_data_tab, text="初期数据")
        self.initial_data_tab_index = notebook.index(initial_data_tab)
        notebook.tab(self.initial_data_tab_index, state="disabled")
        
        initial_button_frame = ttk.Frame(initial_data_tab)
        initial_button_frame.pack(fill=tk.X, pady=(5, 5))
        
        extract_initial_report_button = ttk.Button(
            initial_button_frame,
            text="提取初期报表",
            command=self.extract_initial_report,
            state=tk.DISABLED,
            width=12
        )
        extract_initial_report_button.pack(side=tk.LEFT, padx=(0, 2))
        
        paste_button = ttk.Button(
            initial_button_frame,
            text="粘贴初期报表",
            command=self.paste_initial_report,
            state=tk.DISABLED,
            width=12
        )
        paste_button.pack(side=tk.LEFT)
        
        production_tool_tab = ttk.Frame(notebook)
        notebook.add(production_tool_tab, text="排产工具")
        self.production_tool_tab_index = notebook.index(production_tool_tab)
        notebook.tab(self.production_tool_tab_index, state="disabled")
        
        report_download_tab = ttk.Frame(notebook)
        notebook.add(report_download_tab, text="报表下载")
        self.report_download_tab_index = notebook.index(report_download_tab)
        notebook.tab(self.report_download_tab_index, state="disabled")
        
        submit_decision_tab = ttk.Frame(notebook)
        notebook.add(submit_decision_tab, text="提交决策")
        self.submit_decision_tab_index = notebook.index(submit_decision_tab)
        notebook.tab(self.submit_decision_tab_index, state="disabled")
        
        notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
        standby_frame = ttk.LabelFrame(column3, text="待用", padding="10")
        standby_frame.pack(fill=tk.BOTH, expand=True)
        
        standby_label = ttk.Label(
            standby_frame,
            text="此区域待用",
            font=("Microsoft YaHei UI", 12),
            foreground="#7f8c8d"
        )
        standby_label.pack(expand=True)
        
        self.login_button = login_button
        self.stop_button = stop_button
        self.enter_game_button = enter_game_button
        self.copy_button = copy_button
        self.import_button = import_button
        self.extract_button = extract_button
        self.extract_initial_report_button = extract_initial_report_button
        self.paste_button = paste_button
        
        self.log_text = scrolledtext.ScrolledText(
            None,
            width=80,
            height=20,
            font=("Consolas", 9),
            wrap=tk.WORD
        )
    
    def log(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.logger.info(message)
    
    def open_file(self, file_path_var):
        from tkinter import filedialog
        file_path = filedialog.askopenfilename(
            title="选择文件",
            filetypes=[
                ("Excel文件", "*.xlsx;*.xls;*.xlsm;*.xlsb;*.csv"),
                ("所有文件", "*.*")
            ]
        )
        if file_path:
            file_path_var.set(file_path)
            self.excel_file_path = file_path
            self.log(f"[文件] 已选择文件: {file_path}")
            
            try:
                file_ext = file_path.lower().split('.')[-1]
                content = ""
                
                if file_ext in ['xlsx', 'xls', 'xlsm', 'xlsb', 'csv']:
                    import openpyxl
                    keep_vba = (file_ext == 'xlsm')
                    wb = openpyxl.load_workbook(file_path, read_only=True, keep_vba=keep_vba)
                    ws = wb.active
                    
                    for row in ws.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                            content += row_text + "\n"
                    
                    wb.close()
                    self.log(f"[文件] 已加载Excel文件: {file_path}")
                    
                    has_rules = False
                    for item in self.rules_table.get_children():
                        item_values = self.rules_table.item(item, "values")
                        if item_values and len(item_values) > 1 and item_values[1]:
                            has_rules = True
                            break
                    
                    if has_rules:
                        self.import_button.config(state=tk.NORMAL)
                        self.extract_button.config(state=tk.NORMAL)
                        self.paste_button.config(state=tk.NORMAL)
                        self.log(f"[文件] 规则已复制，已启用导入、提取和粘贴按钮")
                    else:
                        self.import_button.config(state=tk.DISABLED)
                        self.extract_button.config(state=tk.DISABLED)
                        self.paste_button.config(state=tk.DISABLED)
                        self.log(f"[文件] 规则未复制，按钮保持禁用状态")
                    
                else:
                    messagebox.showerror("文件格式错误", f"不支持的文件格式: {file_ext}\n\n仅支持Excel文件格式：\n- XLSX\n- XLS\n- XLSM\n- XLSB\n- CSV")
                    self.log(f"[错误] 不支持的文件格式: {file_ext}")
                    return
                            
            except Exception as e:
                self.log(f"[错误] 读取文件失败: {e}")
                messagebox.showerror("错误", f"无法读取文件：\n\n{e}")
    
    def update_status(self, status, color="black"):
        self.status.set(status)
        self.log(f"[状态] {status}")
    
    def clear_inputs(self):
        self.username.set("")
        self.password.set("")
        
        self.games_listbox.delete(0, tk.END)
        self.games = []
        self.enter_game_button.config(state=tk.DISABLED)
        self.copy_button.config(state=tk.DISABLED)
        self.import_button.config(state=tk.DISABLED)
        self.extract_button.config(state=tk.DISABLED)
        self.paste_button.config(state=tk.DISABLED)
        
        self.update_status("已清空输入框和赛事列表")
        self.log("[操作] 清空了输入框和赛事列表")
    
    def clear_rules_table(self):
        for item in self.rules_table.get_children():
            self.rules_table.set(item, "col2", "")
        self.log("[操作] 清空了规则详情表格的值")
    
    def toggle_password_visibility(self):
        show = self.show_password.get()
        self.password_entry.config(show="" if show else "*")
        self.log(f"[密码] 密码显示模式: {'可见' if show else '隐藏'}")
    
    def bring_to_front(self):
        try:
            self.root.lift()
            self.root.focus_force()
            self.root.attributes('-topmost', True)
            self.root.after(100, lambda: self.root.attributes('-topmost', False))
        except Exception as e:
            self.log(f"[错误] 无法提升界面: {e}")
    
    def show_log_dialog(self):
        log_window = tk.Toplevel(self.root)
        log_window.title("日志")
        log_window.geometry("800x600")
        
        log_frame = ttk.Frame(log_window, padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        log_text = scrolledtext.ScrolledText(
            log_frame,
            width=80,
            height=25,
            font=("Consolas", 9),
            wrap=tk.WORD
        )
        log_text.pack(fill=tk.BOTH, expand=True)
        
        log_text.insert(tk.END, self.log_text.get("1.0", tk.END))
        log_text.config(state=tk.DISABLED)
        
        button_frame = ttk.Frame(log_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        close_button = ttk.Button(
            button_frame,
            text="关闭",
            command=log_window.destroy,
            width=15
        )
        close_button.pack(side=tk.RIGHT)
        
        clear_log_button = ttk.Button(
            button_frame,
            text="清空日志",
            command=lambda: self.clear_log(log_text),
            width=15
        )
        clear_log_button.pack(side=tk.RIGHT, padx=(0, 10))
    
    def clear_log(self, log_text_widget=None):
        self.log_text.delete("1.0", tk.END)
        if log_text_widget:
            log_text_widget.config(state=tk.NORMAL)
            log_text_widget.delete("1.0", tk.END)
            log_text_widget.config(state=tk.DISABLED)
    
    def load_games(self):
        try:
            
            if not self.page_handler:
                self.log("[错误] 页面处理器未初始化")
                return
            
            self.game_extractor = GameExtractor(self.page_handler)
            
            games = self.game_extractor.extract_games_from_table()
            
            if not games:
                games = self.game_extractor.extract_games_with_links()
            
            self.games = games
            
            self.games_listbox.delete(0, tk.END)
            
            for i, game in enumerate(games):
                game_id = game.get('比赛ID', '')
                game_name = game.get('比赛名称', f'赛事 {i+1}')
                create_date = game.get('创建日期', '')
                game_status = game.get('比赛状态', '')
                game_desc = game.get('比赛描述', '')
                
                display_text = f"{game_id}：{game_name}，{create_date}，{game_status}，{game_desc}"
                
                self.games_listbox.insert(tk.END, display_text)
            
            self.update_status(f"找到 {len(games)} 个赛事", color="green")
            
            if games:
                self.enter_game_button.config(state=tk.NORMAL)
            else:
                self.enter_game_button.config(state=tk.DISABLED)
                
        except Exception as e:
            self.log(f"[错误] 加载赛事列表失败: {e}")
            self.update_status("加载赛事列表失败", color="red")
    
    def enter_game(self):
        try:
            if not self.page_handler:
                messagebox.showerror("错误", "浏览器未启动，请先登录")
                self.update_status("浏览器未启动", color="red")
                self.log("[错误] page_handler为None")
                return
                
            selection = self.games_listbox.curselection()
            if not selection:
                messagebox.showwarning("提示", "请先选择一个赛事")
                return
            
            index = selection[0]
            if index >= len(self.games):
                messagebox.showerror("错误", "选择的赛事无效")
                return
            
            game = self.games[index]
            game_url = game.get('url')
            
            if not game_url:
                messagebox.showerror("错误", "该赛事没有有效的链接")
                return
            
            game_id = game.get('比赛ID', '')
            game_name = game.get('比赛名称', 'Unknown')
            self.current_game_id = game_id
            
            if game_url.startswith('/'):
                game_url = f"https://www.ibizsim.cn{game_url}"
            
            self.update_status(f"正在进入赛事: {game_id} - {game_name}", color="blue")
            
            self.playwright_queue.append(('navigate', game_url, game_id, game_name))
            self.log(f"[队列] 已添加导航请求到Playwright队列")
            
            if not self.playwright_thread or not self.playwright_thread.is_alive():
                self.log(f"[Playwright] 启动Playwright操作线程")
                self.playwright_running = True
                self.playwright_thread = threading.Thread(target=self.playwright_operation_loop)
                self.playwright_thread.daemon = True
                self.playwright_thread.start()
                
        except Exception as e:
            self.log(f"[错误] 进入赛事失败: {e}")
            self.update_status("进入赛事失败", color="red")
            messagebox.showerror("错误", f"进入赛事失败：\n\n{e}")
    
    def copy_rules(self):
        try:
            if not self.page_handler:
                messagebox.showerror("错误", "浏览器未启动，请先登录")
                self.update_status("浏览器未启动", color="red")
                self.log("[错误] page_handler为None")
                return
                
            selection = self.games_listbox.curselection()
            if not selection:
                messagebox.showwarning("提示", "请先选择一个赛事")
                return
            
            index = selection[0]
            if index >= len(self.games):
                messagebox.showerror("错误", "选择的赛事无效")
                return
            
            game = self.games[index]
            game_url = game.get('url')
            
            if not game_url:
                messagebox.showerror("错误", "该赛事没有有效的链接")
                return
            
            game_id = game.get('比赛ID', '')
            game_name = game.get('比赛名称', 'Unknown')
            
            rules_url = game_url.replace('/welcome?', '/rules?')
            
            self.log(f"[规则] 正在跳转到规则页面: {game_id} - {game_name}")
            self.log(f"[规则] 原始URL: {game_url}")
            self.log(f"[规则] 规则URL: {rules_url}")
            
            if rules_url.startswith('/'):
                rules_url = f"https://www.ibizsim.cn{rules_url}"
            
            self.log(f"[规则] 转换后URL: {rules_url}")
            self.update_status(f"正在跳转到规则页面: {game_id} - {game_name}", color="blue")
            
            self.playwright_queue.append(('navigate', rules_url, game_id, game_name))
            self.log(f"[队列] 已添加规则导航请求到Playwright队列")
            
            if not self.playwright_thread or not self.playwright_thread.is_alive():
                self.log(f"[Playwright] 启动Playwright操作线程")
                self.playwright_running = True
                self.playwright_thread = threading.Thread(target=self.playwright_operation_loop)
                self.playwright_thread.daemon = True
                self.playwright_thread.start()
                
        except Exception as e:
            self.log(f"[错误] 跳转到规则页面失败: {e}")
            self.update_status("跳转到规则页面失败", color="red")
            messagebox.showerror("错误", f"跳转到规则页面失败：\n\n{e}")
    
    def import_rules(self):
        try:
            if not self.excel_file_path:
                messagebox.showwarning("提示", "请先打开Excel文件")
                self.log(f"[导入] 未打开Excel文件")
                return
            
            self.log(f"[导入] 开始导入规则到Excel文件: {self.excel_file_path}")
            self.update_status("正在导入规则到Excel...", color="blue")
            
            import openpyxl
            file_ext = self.excel_file_path.lower().split('.')[-1]
            keep_vba = (file_ext == 'xlsm')
            
            try:
                wb = openpyxl.load_workbook(self.excel_file_path, keep_vba=keep_vba)
                self.log(f"[导入] 成功打开Excel文件")
            except PermissionError as e:
                self.log(f"[导入] 文件权限错误: {e}")
                messagebox.showerror("错误", f"无法打开Excel文件：\n\n文件可能正在被其他程序使用，请关闭该文件后重试。")
                self.update_status("导入规则到Excel失败", color="red")
                return
            except Exception as e:
                self.log(f"[导入] 打开Excel文件失败: {e}")
                messagebox.showerror("错误", f"打开Excel文件失败：\n\n{e}")
                self.update_status("导入规则到Excel失败", color="red")
                return
            
            ws = None
            for sheet in wb.sheetnames:
                if "初始化表" in sheet:
                    ws = wb[sheet]
                    break
            
            if not ws:
                messagebox.showerror("错误", "Excel文件中未找到'初始化表'工作表")
                self.log(f"[导入] 未找到'初始化表'工作表")
                wb.close()
                return
            
            parameters_to_import = [
                "当期可运输比例",
                "公司总数",
                "公司序号",
                "原材料库存费用",
                "购机费用",
                "原材料固定运费",
                "原材料变动运费",
                "原材料可用比例",
                "维修费",
                "新员工培训费",
                "安置费",
                "基本工资",
                "一加特殊工资",
                "二班正班工资",
                "二加特殊工资",
                "废品系数",
                "最高工资系数",
                "最低资金额度",
                "贷款利息",
                "国债利息",
                "债券利息",
                "税收比例",
                "减税比例",
                "资金有效性",
                "本期利润",
                "市场份额",
                "累计分红",
                "累计缴税",
                "净资产",
                "人均利润率",
                "资本利润率"
            ]
            
            import_count = 0
            for param in parameters_to_import:
                value = ""
                for item in self.rules_table.get_children():
                    item_values = self.rules_table.item(item, "values")
                    if item_values and item_values[0] == param:
                        value = item_values[1] if len(item_values) > 1 else ""
                        break
                
                if value:
                    if param == "当期可运输比例":
                        ws['B4'] = value
                    elif param == "公司总数":
                        ws['B5'] = value
                    elif param == "公司序号":
                        ws['B6'] = value
                    elif param == "原材料库存费用":
                        ws['B7'] = value
                    elif param == "购机费用":
                        ws['B8'] = value
                    elif param == "原材料固定运费":
                        ws['B9'] = value
                    elif param == "原材料变动运费":
                        ws['B10'] = value
                    elif param == "原材料可用比例":
                        ws['B11'] = value
                    elif param == "维修费":
                        ws['B12'] = value
                    elif param == "新员工培训费":
                        ws['B13'] = value
                    elif param == "安置费":
                        ws['B14'] = value
                    elif param == "基本工资":
                        ws['B15'] = value
                    elif param == "一加特殊工资":
                        ws['B16'] = value
                    elif param == "二班正班工资":
                        ws['B17'] = value
                    elif param == "二加特殊工资":
                        ws['B18'] = value
                    elif param == "废品系数":
                        ws['B19'] = value
                    elif param == "最高工资系数":
                        ws['B20'] = value
                    elif param == "最低资金额度":
                        ws['B21'] = value
                    elif param == "贷款利息":
                        ws['B22'] = value
                    elif param == "国债利息":
                        ws['B23'] = value
                    elif param == "债券利息":
                        ws['B24'] = value
                    elif param == "税收比例":
                        ws['B25'] = value
                    elif param == "减税比例":
                        ws['B26'] = value
                    elif param == "资金有效性":
                        ws['B27'] = value
                    elif param == "本期利润":
                        ws['B30'] = value
                    elif param == "市场份额":
                        ws['B31'] = value
                    elif param == "累计分红":
                        ws['B32'] = value
                    elif param == "累计缴税":
                        ws['B33'] = value
                    elif param == "净资产":
                        ws['B34'] = value
                    elif param == "人均利润率":
                        ws['B35'] = value
                    elif param == "资本利润率":
                        ws['B36'] = value
                    import_count += 1
            
            product_inventory_count = 0
            management_count = 0
            for item in self.product_table.get_children():
                item_values = self.product_table.item(item, "values")
                if item_values and len(item_values) >= 2 and item_values[1]:
                    product_name = item_values[0]
                    inventory_cost = item_values[1]
                    if "产品1" in product_name:
                        ws['E5'] = inventory_cost
                        product_inventory_count += 1
                    elif "产品2" in product_name:
                        ws['E6'] = inventory_cost
                        product_inventory_count += 1
                    elif "产品3" in product_name:
                        ws['E7'] = inventory_cost
                        product_inventory_count += 1
                    elif "产品4" in product_name:
                        ws['E8'] = inventory_cost
                        product_inventory_count += 1
                
                if item_values and len(item_values) >= 5:
                    product_name = item_values[0]
                    if "产品1" in product_name:
                        if item_values[3]:
                            ws['H25'] = item_values[3]
                            management_count += 1
                        if item_values[4]:
                            ws['I25'] = item_values[4]
                            management_count += 1
                    elif "产品2" in product_name:
                        if item_values[3]:
                            ws['H26'] = item_values[3]
                            management_count += 1
                        if item_values[4]:
                            ws['I26'] = item_values[4]
                            management_count += 1
                    elif "产品3" in product_name:
                        if item_values[3]:
                            ws['H27'] = item_values[3]
                            management_count += 1
                        if item_values[4]:
                            ws['I27'] = item_values[4]
                            management_count += 1
                    elif "产品4" in product_name:
                        if item_values[3]:
                            ws['H28'] = item_values[3]
                            management_count += 1
                        if item_values[4]:
                            ws['I28'] = item_values[4]
                            management_count += 1
            
            import_count += product_inventory_count
            
            conversion_count = 0
            for item in self.conversion_table.get_children():
                item_values = self.conversion_table.item(item, "values")
                if item_values and len(item_values) >= 5:
                    market_name = item_values[0]
                    for col_idx in range(1, 5):
                        if item_values[col_idx]:
                            if "市场1" in market_name:
                                ws.cell(row=11, column=5 + col_idx - 1, value=item_values[col_idx])
                                conversion_count += 1
                            elif "市场2" in market_name:
                                ws.cell(row=12, column=5 + col_idx - 1, value=item_values[col_idx])
                                conversion_count += 1
                            elif "市场3" in market_name:
                                ws.cell(row=13, column=5 + col_idx - 1, value=item_values[col_idx])
                                conversion_count += 1
                            elif "市场4" in market_name:
                                ws.cell(row=14, column=5 + col_idx - 1, value=item_values[col_idx])
                                conversion_count += 1
            
            import_count += conversion_count
            import_count += management_count
            
            production_count = 0
            for item in self.production_table.get_children():
                item_values = self.production_table.item(item, "values")
                if item_values and len(item_values) >= 5:
                    item_name = item_values[0]
                    for col_idx in range(1, 5):
                        if item_values[col_idx]:
                            if "机器" in item_name:
                                ws.cell(row=17, column=5 + col_idx - 1, value=item_values[col_idx])
                                production_count += 1
                            elif "人力" in item_name:
                                ws.cell(row=18, column=5 + col_idx - 1, value=item_values[col_idx])
                                production_count += 1
                            elif "原材料" in item_name:
                                ws.cell(row=19, column=5 + col_idx - 1, value=item_values[col_idx])
                                production_count += 1
            
            import_count += production_count
            
            discount_count = 0
            for item in self.discount_table.get_children():
                item_values = self.discount_table.item(item, "values")
                if item_values and len(item_values) >= 2 and item_values[1]:
                    discount_count += 1
                    if discount_count == 1:
                        ws['D26'] = item_values[0]
                        ws['E26'] = item_values[1]
                    elif discount_count == 2:
                        ws['D27'] = item_values[0]
                        ws['E27'] = item_values[1]
                    elif discount_count == 3:
                        ws['D28'] = item_values[0]
                        ws['E28'] = item_values[1]
            
            import_count += discount_count
            
            rd_count = 0
            for item in self.grade_table.get_children():
                item_values = self.grade_table.item(item, "values")
                if item_values and len(item_values) >= 6:
                    product_name = item_values[0]
                    for col_idx in range(1, 6):
                        if item_values[col_idx]:
                            if "产品1" in product_name:
                                ws.cell(row=31, column=5 + col_idx - 1, value=item_values[col_idx])
                                rd_count += 1
                            elif "产品2" in product_name:
                                ws.cell(row=32, column=5 + col_idx - 1, value=item_values[col_idx])
                                rd_count += 1
                            elif "产品3" in product_name:
                                ws.cell(row=33, column=5 + col_idx - 1, value=item_values[col_idx])
                                rd_count += 1
                            elif "产品4" in product_name:
                                ws.cell(row=34, column=5 + col_idx - 1, value=item_values[col_idx])
                                rd_count += 1
                            elif "产品5" in product_name:
                                ws.cell(row=35, column=5 + col_idx - 1, value=item_values[col_idx])
                                rd_count += 1
            
            import_count += rd_count
            
            fixed_shipping_count = 0
            shipping_items = self.shipping_table.get_children()
            
            if len(shipping_items) >= 4:
                product1_item = shipping_items[0]
                product1_values = list(self.shipping_table.item(product1_item, "values"))
                if len(product1_values) >= 5:
                    if product1_values[1]:
                        ws['L5'] = product1_values[1]
                        fixed_shipping_count += 1
                    if product1_values[2]:
                        ws['M5'] = product1_values[2]
                        fixed_shipping_count += 1
                    if product1_values[3]:
                        ws['N5'] = product1_values[3]
                        fixed_shipping_count += 1
                    if product1_values[4]:
                        ws['O5'] = product1_values[4]
                        fixed_shipping_count += 1
                
                product2_item = shipping_items[1]
                product2_values = list(self.shipping_table.item(product2_item, "values"))
                if len(product2_values) >= 5:
                    if product2_values[1]:
                        ws['P5'] = product2_values[1]
                        fixed_shipping_count += 1
                    if product2_values[2]:
                        ws['Q5'] = product2_values[2]
                        fixed_shipping_count += 1
                    if product2_values[3]:
                        ws['R5'] = product2_values[3]
                        fixed_shipping_count += 1
                    if product2_values[4]:
                        ws['S5'] = product2_values[4]
                        fixed_shipping_count += 1
                
                product3_item = shipping_items[2]
                product3_values = list(self.shipping_table.item(product3_item, "values"))
                if len(product3_values) >= 5:
                    if product3_values[1]:
                        ws['L6'] = product3_values[1]
                        fixed_shipping_count += 1
                    if product3_values[2]:
                        ws['M6'] = product3_values[2]
                        fixed_shipping_count += 1
                    if product3_values[3]:
                        ws['N6'] = product3_values[3]
                        fixed_shipping_count += 1
                    if product3_values[4]:
                        ws['O6'] = product3_values[4]
                        fixed_shipping_count += 1
                
                product4_item = shipping_items[3]
                product4_values = list(self.shipping_table.item(product4_item, "values"))
                if len(product4_values) >= 5:
                    if product4_values[1]:
                        ws['P6'] = product4_values[1]
                        fixed_shipping_count += 1
                    if product4_values[2]:
                        ws['Q6'] = product4_values[2]
                        fixed_shipping_count += 1
                    if product4_values[3]:
                        ws['R6'] = product4_values[3]
                        fixed_shipping_count += 1
                    if product4_values[4]:
                        ws['S6'] = product4_values[4]
                        fixed_shipping_count += 1
            
            import_count += fixed_shipping_count
            
            variable_shipping_count = 0
            shipping3_items = self.shipping_table3.get_children()
            
            if len(shipping3_items) >= 4:
                product1_item = shipping3_items[0]
                product1_values = list(self.shipping_table3.item(product1_item, "values"))
                if len(product1_values) >= 5:
                    if product1_values[1]:
                        ws['L9'] = product1_values[1]
                        variable_shipping_count += 1
                    if product1_values[2]:
                        ws['M9'] = product1_values[2]
                        variable_shipping_count += 1
                    if product1_values[3]:
                        ws['N9'] = product1_values[3]
                        variable_shipping_count += 1
                    if product1_values[4]:
                        ws['O9'] = product1_values[4]
                        variable_shipping_count += 1
                
                product2_item = shipping3_items[1]
                product2_values = list(self.shipping_table3.item(product2_item, "values"))
                if len(product2_values) >= 5:
                    if product2_values[1]:
                        ws['P9'] = product2_values[1]
                        variable_shipping_count += 1
                    if product2_values[2]:
                        ws['Q9'] = product2_values[2]
                        variable_shipping_count += 1
                    if product2_values[3]:
                        ws['R9'] = product2_values[3]
                        variable_shipping_count += 1
                    if product2_values[4]:
                        ws['S9'] = product2_values[4]
                        variable_shipping_count += 1
                
                product3_item = shipping3_items[2]
                product3_values = list(self.shipping_table3.item(product3_item, "values"))
                if len(product3_values) >= 5:
                    if product3_values[1]:
                        ws['L10'] = product3_values[1]
                        variable_shipping_count += 1
                    if product3_values[2]:
                        ws['M10'] = product3_values[2]
                        variable_shipping_count += 1
                    if product3_values[3]:
                        ws['N10'] = product3_values[3]
                        variable_shipping_count += 1
                    if product3_values[4]:
                        ws['O10'] = product3_values[4]
                        variable_shipping_count += 1
                
                product4_item = shipping3_items[3]
                product4_values = list(self.shipping_table3.item(product4_item, "values"))
                if len(product4_values) >= 5:
                    if product4_values[1]:
                        ws['P10'] = product4_values[1]
                        variable_shipping_count += 1
                    if product4_values[2]:
                        ws['Q10'] = product4_values[2]
                        variable_shipping_count += 1
                    if product4_values[3]:
                        ws['R10'] = product4_values[3]
                        variable_shipping_count += 1
                    if product4_values[4]:
                        ws['S10'] = product4_values[4]
                        variable_shipping_count += 1
            
            import_count += variable_shipping_count
            wb.save(self.excel_file_path)
            wb.close()
            
            self.update_status(f"已导入 {import_count} 个参数值到Excel", color="green")
            self.log(f"[导入] 成功导入 {import_count} 个参数值到Excel文件")
            messagebox.showinfo("成功", f"已成功导入 {import_count} 个参数值到Excel文件\n\n包括31个规则参数、{product_inventory_count}个成品库存费、{management_count}个管理费、{conversion_count}个订货转化比例、{production_count}个产品生产消耗、{discount_count}个原材料折扣、{rd_count}个研发投入、{fixed_shipping_count}个固定运费和{variable_shipping_count}个变动运费")
            
            self.notebook.tab(self.salary_coefficient_tab_index, state="normal")
            self.notebook.tab(self.initial_data_tab_index, state="normal")
            self.notebook.tab(self.production_tool_tab_index, state="normal")
            self.notebook.tab(self.report_download_tab_index, state="normal")
            self.notebook.tab(self.submit_decision_tab_index, state="normal")
            self.log("[导入] 已启用工资系数、初期数据、排产工具、报表下载、提交决策功能")
            
            try:
                import os
                os.startfile(self.excel_file_path)
                self.log(f"[文件] 已用系统默认程序打开: {self.excel_file_path}")
                
                self.root.focus_force()
                self.root.after(500, self.root.focus_set)
            except Exception as e:
                self.log(f"[文件] 用系统程序打开失败: {e}")
            
        except Exception as e:
            self.log(f"[错误] 导入规则到Excel失败: {e}")
            self.update_status("导入规则到Excel失败", color="red")
            messagebox.showerror("错误", f"导入规则到Excel失败：\n\n{e}")
    
    def on_tab_changed(self, event):
        try:
            current_tab_index = self.notebook.index(self.notebook.select())
            
            if current_tab_index in [self.salary_coefficient_tab_index, self.initial_data_tab_index]:
                if not self.current_game_id or not self.current_team_id:
                    self.log("[Tab切换] 缺少game_id或team_id，无法访问报表页面")
                    return
                
                url = f"https://www.ibizsim.cn/games/private_report?gameid={self.current_game_id}&teamid={self.current_team_id}"
                self.log(f"[Tab切换] 访问报表页面: {url}")
                
                if not self.playwright_thread or not self.playwright_thread.is_alive():
                    self.log("[错误] Playwright线程未运行")
                    return
                
                self.playwright_queue.append(('extract_periodid', url))
                    
        except Exception as e:
            self.log(f"[错误] Tab切换处理失败: {e}")
    
    def extract_quality_rates(self):
        try:
            if not self.page_handler:
                messagebox.showerror("错误", "浏览器未启动，请先登录")
                self.update_status("浏览器未启动", color="red")
                self.log("[错误] page_handler为None")
                return
                
            self.log("[正品率] 开始提取前八期正品率...")
            self.update_status("正在提取前八期正品率", color="blue")
            
            messagebox.showinfo("提示", "前八期正品率提取功能开发中...")
            
        except Exception as e:
            self.log(f"[错误] 提取前八期正品率失败: {e}")
            self.update_status("提取前八期正品率失败", color="red")
            messagebox.showerror("错误", f"提取前八期正品率失败：\n\n{e}")
    
    def extract_initial_report(self):
        try:
            if not self.page_handler:
                messagebox.showerror("错误", "浏览器未启动，请先登录")
                self.update_status("浏览器未启动", color="red")
                self.log("[错误] page_handler为None")
                return
            
            if not self.period_8_url:
                messagebox.showerror("错误", "第8期报表URL未提取，请先切换到初期数据tab")
                self.log("[错误] 第8期报表URL未提取")
                return
            
            self.log(f"[报表] 正在跳转到第8期报表页面...")
            self.update_status("正在跳转到第8期报表页面", color="blue")
            
            if not self.playwright_thread or not self.playwright_thread.is_alive():
                self.log("[错误] Playwright线程未运行")
                return
            
            self.playwright_queue.append(('navigate_period8', self.period_8_url))
            
        except Exception as e:
            self.log(f"[错误] 提取初期报表失败: {e}")
            self.update_status("提取初期报表失败", color="red")
            messagebox.showerror("错误", f"提取初期报表失败：\n\n{e}")
    
    def paste_initial_report(self):
        try:
            if not self.page_handler:
                messagebox.showerror("错误", "浏览器未启动，请先登录")
                self.update_status("浏览器未启动", color="red")
                self.log("[错误] page_handler为None")
                return
                
            self.log("[报表] 开始粘贴初期报表...")
            self.update_status("正在粘贴初期报表", color="blue")
            
            messagebox.showinfo("提示", "初期报表粘贴功能开发中...")
            
        except Exception as e:
            self.log(f"[错误] 粘贴初期报表失败: {e}")
            self.update_status("粘贴初期报表失败", color="red")
            messagebox.showerror("错误", f"粘贴初期报表失败：\n\n{e}")
    
    def extract_rules_parameters_in_thread(self):
        try:
            if not self.page_handler:
                self.log("[错误] 页面处理器未初始化")
                return
            
            
            page = self.page_handler.get_page()
            if not page:
                self.log("[错误] 无法获取页面对象")
                return
            
            parameters = [
                "当期可运输比例",
                "公司总数",
                "公司序号",
                "原材料库存费用",
                "购机费用",
                "原材料固定运费",
                "原材料变动运费",
                "原材料可用比例",
                "维修费",
                "新员工培训费",
                "安置费",
                "基本工资",
                "一加特殊工资",
                "二班正班工资",
                "二加特殊工资",
                "废品系数",
                "最高工资系数",
                "最低资金额度",
                "贷款利息",
                "国债利息",
                "债券利息",
                "税收比例",
                "减税比例",
                "资金有效性",
                "本期利润",
                "市场份额",
                "累计分红",
                "累计缴税",
                "净资产",
                "人均利润率",
                "资本利润率"
            ]
            
            param_values = {}
            
            for param in parameters:
                try:
                    value = ""
                    
                    if param == "当期可运输比例":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"本期产品的(\d+\.?\d*%)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "公司总数":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"参加本次模拟的有(\d+)个组"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "公司序号":
                        try:
                            if self.team_name:
                                page_content = page.content()
                                import re
                                table_pattern = r'<table[^>]*class="table[^"]*"[^>]*>.*?<tbody>(.*?)</tbody>'
                                table_match = re.search(table_pattern, page_content, re.DOTALL)
                                if table_match:
                                    table_content = table_match.group(1)
                                    self.log(f"[调试] 表格内容长度: {len(table_content)}")
                                    
                                    all_tr_matches = list(re.finditer(r'<tr[^>]*>', table_content))
                                    self.log(f"[调试] 找到 {len(all_tr_matches)} 个 <tr> 标签")
                                    
                                    company_number = 0
                                    for i, tr_match in enumerate(all_tr_matches):
                                        tr_start = tr_match.start()
                                        tr_end = tr_match.end()
                                        next_tr_start = all_tr_matches[i+1].start() if i+1 < len(all_tr_matches) else len(table_content)
                                        tr_content = table_content[tr_start:next_tr_start]
                                        
                                        if self.team_name in tr_content:
                                            company_number = i - 1
                                            self.log(f"[调试] 第{i+1}行包含队伍名称: {self.team_name}")
                                            self.log(f"[调试] 确定公司序号: {company_number} (跳过前2个表头行)")
                                            break
                                    
                                    if company_number > 0:
                                        value = str(company_number)
                        except Exception as e:
                            import traceback
                            self.log(f"[调试] 错误详情: {traceback.format_exc()}")
                    elif param == "原材料库存费用":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"原材料每期(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "购机费用":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"机器价格为([\d,\.]+)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                raw_value = match.group(1)
                                if '.' in raw_value:
                                    raw_value = raw_value.split('.')[0]
                                value = raw_value.replace(',', '')
                        except Exception as e:
                            pass
                    elif param == "原材料固定运费":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"固定费用为(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "原材料变动运费":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"变动费用为(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "原材料可用比例":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"至多有(\d+\.?\d*%)可以"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "维修费":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"维修费为(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                raw_value = match.group(1)
                                if '.' in raw_value:
                                    raw_value = raw_value.split('.')[0]
                                value = raw_value.replace(',', '')
                        except Exception as e:
                            pass
                    elif param == "新员工培训费":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"培训费为(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                raw_value = match.group(1)
                                if '.' in raw_value:
                                    raw_value = raw_value.split('.')[0]
                                value = raw_value.replace(',', '')
                        except Exception as e:
                            pass
                    elif param == "安置费":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"安置费(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                raw_value = match.group(1)
                                if '.' in raw_value:
                                    raw_value = raw_value.split('.')[0]
                                value = raw_value.replace(',', '')
                        except Exception as e:
                            pass
                    elif param == "基本工资":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"第一班正班：(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                raw_value = match.group(1)
                                if '.' in raw_value:
                                    raw_value = raw_value.split('.')[0]
                                value = raw_value.replace(',', '')
                        except Exception as e:
                            pass
                    elif param == "一加特殊工资":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"第一班加班：(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                raw_value = match.group(1)
                                if '.' in raw_value:
                                    raw_value = raw_value.split('.')[0]
                                value = raw_value.replace(',', '')
                        except Exception as e:
                            pass
                    elif param == "二班正班工资":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"第二班正班：(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                raw_value = match.group(1)
                                if '.' in raw_value:
                                    raw_value = raw_value.split('.')[0]
                                value = raw_value.replace(',', '')
                        except Exception as e:
                            pass
                    elif param == "二加特殊工资":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"第二班加班：(\d+\.?\d*)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                raw_value = match.group(1)
                                if '.' in raw_value:
                                    raw_value = raw_value.split('.')[0]
                                value = raw_value.replace(',', '')
                        except Exception as e:
                            pass
                    elif param == "废品系数":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"废品系数[为是]?(\d+\.?\d*)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                            else:
                                value = "1"
                        except Exception as e:
                            value = "1"
                    elif param == "最高工资系数":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"最高工资系数[为是]?(\d+\.?\d*)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                            else:
                                value = "1.4"
                        except Exception as e:
                            value = "1.4"
                    elif param == "最低资金额度":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"至少应有([\d,\.]+)元"
                            match = re.search(pattern, rule_content)
                            if match:
                                raw_value = match.group(1)
                                value = raw_value.replace(',', '')
                        except Exception as e:
                            pass
                    elif param == "贷款利息":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"银行贷款的本利在本期末偿还，年利率为(\d+\.?\d*%)（"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "国债利息":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"公司每期都可以买国债，年利率为(\d+\.?\d*%)。"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "债券利息":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"债券的年利率为(\d+\.?\d*%)。"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "税收比例":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"净收益的(\d+\.?\d*%)，"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "减税比例":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"亏损额的(\d+\.?\d*%)在"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "资金有效性":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"分红按(\d+\.?\d*%)的"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "本期利润":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"各项指标的权重分别为[：:]\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(1)
                        except Exception as e:
                            pass
                    elif param == "市场份额":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"各项指标的权重分别为[：:]\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(2)
                        except Exception as e:
                            pass
                    elif param == "累计分红":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"各项指标的权重分别为[：:]\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(3)
                        except Exception as e:
                            pass
                    elif param == "累计缴税":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"各项指标的权重分别为[：:]\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(4)
                        except Exception as e:
                            pass
                    elif param == "净资产":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"各项指标的权重分别为[：:]\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(5)
                        except Exception as e:
                            pass
                    elif param == "人均利润率":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"各项指标的权重分别为[：:]\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(6)
                        except Exception as e:
                            pass
                    elif param == "资本利润率":
                        try:
                            rule_content = page.locator("#rule").inner_text()
                            import re
                            pattern = r"各项指标的权重分别为[：:]\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)\s*,\s*([\d\.]+)"
                            match = re.search(pattern, rule_content)
                            if match:
                                value = match.group(7)
                        except Exception as e:
                            pass
                    else:
                        selectors = [
                            f"#rule//*[contains(text(), '{param}')]/following-sibling::*[1]",
                            f"#rule//*[contains(text(), '{param}')]/following-sibling::*[2]",
                            f"#rule//*[contains(text(), '{param}')]/parent::*//*[contains(@class, 'value')]",
                            f"#rule//*[contains(text(), '{param}')]/parent::*//*[contains(@class, 'param-value')]",
                            f"#rule//td[contains(text(), '{param}')]/following-sibling::td[1]",
                            f"#rule//td[contains(text(), '{param}')]/following-sibling::td[2]",
                            f"#rule//div[contains(text(), '{param}')]/following-sibling::div[1]",
                            f"#rule//div[contains(text(), '{param}')]/following-sibling::div[2]",
                            f"#rule//span[contains(text(), '{param}')]/following-sibling::span[1]",
                            f"#rule//span[contains(text(), '{param}')]/following-sibling::span[2]",
                            f"#rule//*[contains(text(), '{param}')]/parent::*//td[position()>1]",
                            f"#rule//*[contains(text(), '{param}')]/parent::*//div[position()>1]",
                            f"#rule//*[contains(text(), '{param}')]/parent::*//span[position()>1]",
                            f"#rule//text()[contains(., '{param}')]/parent::*/following-sibling::*[1]",
                            f"#rule//text()[contains(., '{param}')]/parent::*/following-sibling::*[2]",
                            f"#rule//*[contains(text(), '{param}')]/following::*[1]",
                            f"#rule//*[contains(text(), '{param}')]/following::*[2]",
                            f"#rule//*[contains(text(), '{param}')]/ancestor::*/following-sibling::*[1]",
                            f"#rule//*[contains(text(), '{param}')]/ancestor::*/following-sibling::*[2]"
                        ]
                        
                        for selector in selectors:
                            try:
                                elements = page.query_selector_all(selector)
                                if elements:
                                    for elem in elements:
                                        text = elem.text_content().strip()
                                        if text and text != param:
                                            value = text
                                            break
                                    if value:
                                        break
                            except:
                                continue
                    
                    param_values[param] = value
                    
                except Exception as e:
                    param_values[param] = ""
            
            for item in self.rules_table.get_children():
                self.rules_table.delete(item)
            
            for param in parameters:
                value = param_values.get(param, "")
                self.rules_table.insert("", tk.END, values=(param, value))
            
            self.update_status(f"已提取 {len([v for v in param_values.values() if v])} 个参数值", color="green")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'库存费为[：:]\s*'
                end_pattern = r'库存费在每期'
                
                start_match = re.search(start_pattern, page_content)
                end_match = re.search(end_pattern, page_content)
                
                if start_match and end_match:
                    start_pos = start_match.end()
                    end_pos = end_match.start()
                    
                    target_content = page_content[start_pos:end_pos]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_match = re.search(table_pattern, target_content, re.DOTALL)
                    
                    if table_match:
                        tbody_content = table_match.group(1)
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        all_numbers = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                number_pattern = r'(\d+\.?\d*|\d+,?\d*)'
                                numbers = re.findall(number_pattern, clean_text)
                                
                                for num in numbers:
                                    num = num.replace(',', '')
                                    try:
                                        float_num = float(num)
                                        all_numbers.append(num)
                                    except ValueError:
                                        pass
                        
                        if all_numbers:
                            
                            product_items = self.product_table.get_children()
                            for i, item in enumerate(product_items):
                                if i < len(all_numbers):
                                    current_values = list(self.product_table.item(item, "values"))
                                    current_values[1] = all_numbers[i]
                                    self.product_table.item(item, values=current_values)
                    else:
                        self.log(f"[调试] 目标区域前500字符: {target_content[:500]}")
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'具体见下表（单位：元）'
                end_pattern = r'维修费'
                
                start_match = re.search(start_pattern, page_content)
                end_match = re.search(end_pattern, page_content)
                
                if start_match and end_match:
                    start_pos = start_match.end()
                    end_pos = end_match.start()
                    
                    target_content = page_content[start_pos:end_pos]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_match = re.search(table_pattern, target_content, re.DOTALL)
                    
                    if table_match:
                        tbody_content = table_match.group(1)
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        table_data = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            row_numbers = []
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                number_pattern = r'(\d+\.?\d*|\d+,?\d*)'
                                numbers = re.findall(number_pattern, clean_text)
                                
                                for num in numbers:
                                    num = num.replace(',', '')
                                    try:
                                        float_num = float(num)
                                        row_numbers.append(num)
                                    except ValueError:
                                        pass
                            
                            if row_numbers:
                                table_data.append(row_numbers)
                        
                        if table_data:
                            first_col_numbers = []
                            second_col_numbers = []
                            
                            for row in table_data:
                                if len(row) >= 1:
                                    first_col_numbers.append(row[0])
                                if len(row) >= 2:
                                    second_col_numbers.append(row[1])
                            
                            
                            product_items = self.product_table.get_children()
                            for i, item in enumerate(product_items):
                                current_values = list(self.product_table.item(item, "values"))
                                
                                if i < len(first_col_numbers):
                                    current_values[3] = first_col_numbers[i]
                                
                                if i < len(second_col_numbers):
                                    current_values[4] = second_col_numbers[i]
                                
                                self.product_table.item(item, values=current_values)
                    else:
                        self.log(f"[调试] 目标区域前500字符: {target_content[:500]}")
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'<h3>预订</h3>'
                start_match = re.search(start_pattern, page_content)
                
                if start_match:
                    start_pos = start_match.end()
                    
                    target_content = page_content[start_pos:]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_match = re.search(table_pattern, target_content, re.DOTALL)
                    
                    if table_match:
                        table_start_pos = start_pos + table_match.start()
                        table_end_pos = start_pos + table_match.end()
                        
                        tbody_content = table_match.group(1)
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        table_data = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            row_numbers = []
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                number_pattern = r'(\d+\.?\d*%)'
                                numbers = re.findall(number_pattern, clean_text)
                                
                                for num in numbers:
                                    row_numbers.append(num)
                            
                            if row_numbers:
                                table_data.append(row_numbers)
                        
                        if table_data:
                            col_numbers = [[], [], [], []]
                            
                            for row in table_data:
                                for col_idx in range(4):
                                    if col_idx < len(row):
                                        col_numbers[col_idx].append(row[col_idx])
                            
                            
                            market_items = self.conversion_table.get_children()
                            for i, item in enumerate(market_items):
                                current_values = list(self.conversion_table.item(item, "values"))
                                
                                if i < len(col_numbers[0]):
                                    current_values[1] = col_numbers[0][i]
                                
                                if i < len(col_numbers[1]):
                                    current_values[2] = col_numbers[1][i]
                                
                                if i < len(col_numbers[2]):
                                    current_values[3] = col_numbers[2][i]
                                
                                if i < len(col_numbers[3]):
                                    current_values[4] = col_numbers[3][i]
                                
                                self.conversion_table.item(item, values=current_values)
                    else:
                        self.log(f"[调试] 目标区域前500字符: {target_content[:500]}")
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'<h3>生产单个产品所需要的资源</h3>'
                start_match = re.search(start_pattern, page_content)
                
                if start_match:
                    start_pos = start_match.end()
                    
                    target_content = page_content[start_pos:]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_match = re.search(table_pattern, target_content, re.DOTALL)
                    
                    if table_match:
                        table_start_pos = start_pos + table_match.start()
                        table_end_pos = start_pos + table_match.end()
                        
                        tbody_content = table_match.group(1)
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        table_data = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            row_numbers = []
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                number_pattern = r'(\d+\.?\d*)'
                                numbers = re.findall(number_pattern, clean_text)
                                
                                for num in numbers:
                                    num = num.replace(',', '')
                                    try:
                                        float_num = float(num)
                                        row_numbers.append(num)
                                    except ValueError:
                                        pass
                            
                            if row_numbers:
                                table_data.append(row_numbers)
                        
                        if table_data:
                            col_numbers = [[], [], [], []]
                            
                            for row in table_data:
                                for col_idx in range(4):
                                    if col_idx < len(row):
                                        col_numbers[col_idx].append(row[col_idx])
                            
                            
                            production_items = self.production_table.get_children()
                            for i, item in enumerate(production_items):
                                current_values = list(self.production_table.item(item, "values"))
                                
                                if i < len(col_numbers[0]):
                                    current_values[1] = col_numbers[0][i]
                                
                                if i < len(col_numbers[1]):
                                    current_values[2] = col_numbers[1][i]
                                
                                if i < len(col_numbers[2]):
                                    current_values[3] = col_numbers[2][i]
                                
                                if i < len(col_numbers[3]):
                                    current_values[4] = col_numbers[3][i]
                                
                                self.production_table.item(item, values=current_values)
                    else:
                        self.log(f"[调试] 目标区域前500字符: {target_content[:500]}")
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'<h3>研发费用</h3>'
                start_match = re.search(start_pattern, page_content)
                
                if start_match:
                    start_pos = start_match.end()
                    
                    target_content = page_content[start_pos:]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_match = re.search(table_pattern, target_content, re.DOTALL)
                    
                    if table_match:
                        table_start_pos = start_pos + table_match.start()
                        table_end_pos = start_pos + table_match.end()
                        
                        tbody_content = table_match.group(1)
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        table_data = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            row_numbers = []
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                number_pattern = r'(\d+\.?\d*)'
                                numbers = re.findall(number_pattern, clean_text)
                                
                                if numbers:
                                    combined_number = ''.join(numbers)
                                    try:
                                        float_num = float(combined_number)
                                        row_numbers.append(combined_number)
                                    except ValueError:
                                        pass
                            
                            if row_numbers and tr_idx > 0:
                                table_data.append(row_numbers)
                        
                        if table_data:
                            col_numbers = [[], [], [], [], []]
                            
                            for row in table_data:
                                for col_idx in range(5):
                                    if col_idx < len(row):
                                        col_numbers[col_idx].append(row[col_idx])
                            
                            
                            grade_items = self.grade_table.get_children()
                            for i, item in enumerate(grade_items):
                                current_values = list(self.grade_table.item(item, "values"))
                                
                                if i < len(col_numbers[0]):
                                    current_values[1] = col_numbers[0][i]
                                
                                if i < len(col_numbers[1]):
                                    current_values[2] = col_numbers[1][i]
                                
                                if i < len(col_numbers[2]):
                                    current_values[3] = col_numbers[2][i]
                                
                                if i < len(col_numbers[3]):
                                    current_values[4] = col_numbers[3][i]
                                
                                if i < len(col_numbers[4]):
                                    current_values[5] = col_numbers[4][i]
                                
                                self.grade_table.item(item, values=current_values)
                    else:
                        self.log(f"[调试] 目标区域前500字符: {target_content[:500]}")
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'产品运输固定费用（元）'
                start_match = re.search(start_pattern, page_content)
                
                if start_match:
                    start_pos = start_match.end()
                    
                    target_content = page_content[start_pos:]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_match = re.search(table_pattern, target_content, re.DOTALL)
                    
                    if table_match:
                        table_start_pos = start_pos + table_match.start()
                        table_end_pos = start_pos + table_match.end()
                        
                        tbody_content = table_match.group(1)
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        table_data = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            row_numbers = []
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                if clean_text:
                                    number_pattern = r'^[\d,]+\.?\d*$'
                                    if re.match(number_pattern, clean_text):
                                        clean_number = clean_text.replace(',', '')
                                        try:
                                            float_num = float(clean_number)
                                            row_numbers.append(clean_number)
                                        except ValueError:
                                            pass
                            
                            if row_numbers and tr_idx > 0:
                                table_data.append(row_numbers)
                        
                        if table_data:
                            shipping_items = self.shipping_table.get_children()
                            
                            if len(table_data) >= 1 and len(shipping_items) >= 2:
                                row_data = table_data[0]
                                
                                product1_item = shipping_items[0]
                                product1_values = list(self.shipping_table.item(product1_item, "values"))
                                if len(row_data) >= 1:
                                    product1_values[1] = row_data[0]
                                if len(row_data) >= 2:
                                    product1_values[2] = row_data[1]
                                if len(row_data) >= 3:
                                    product1_values[3] = row_data[2]
                                if len(row_data) >= 4:
                                    product1_values[4] = row_data[3]
                                self.shipping_table.item(product1_item, values=product1_values)
                                
                                product2_item = shipping_items[1]
                                product2_values = list(self.shipping_table.item(product2_item, "values"))
                                if len(row_data) >= 5:
                                    product2_values[1] = row_data[4]
                                if len(row_data) >= 6:
                                    product2_values[2] = row_data[5]
                                if len(row_data) >= 7:
                                    product2_values[3] = row_data[6]
                                if len(row_data) >= 8:
                                    product2_values[4] = row_data[7]
                                self.shipping_table.item(product2_item, values=product2_values)
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'产品运输固定费用（元）'
                start_match = re.search(start_pattern, page_content)
                
                if start_match:
                    start_pos = start_match.end()
                    
                    target_content = page_content[start_pos:]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_matches = re.findall(table_pattern, target_content, re.DOTALL)
                    
                    if len(table_matches) >= 2:
                        second_table_match = table_matches[1]
                        tbody_content = second_table_match
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        table_data = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            row_numbers = []
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                if clean_text:
                                    number_pattern = r'^[\d,]+\.?\d*$'
                                    if re.match(number_pattern, clean_text):
                                        clean_number = clean_text.replace(',', '')
                                        try:
                                            float_num = float(clean_number)
                                            row_numbers.append(clean_number)
                                        except ValueError:
                                            pass
                            
                            if row_numbers and tr_idx > 0:
                                table_data.append(row_numbers)
                        
                        if table_data:
                            shipping_items = self.shipping_table.get_children()
                            
                            if len(table_data) >= 1 and len(shipping_items) >= 4:
                                row_data = table_data[0]
                                
                                product3_item = shipping_items[2]
                                product3_values = list(self.shipping_table.item(product3_item, "values"))
                                if len(row_data) >= 1:
                                    product3_values[1] = row_data[0]
                                if len(row_data) >= 2:
                                    product3_values[2] = row_data[1]
                                if len(row_data) >= 3:
                                    product3_values[3] = row_data[2]
                                if len(row_data) >= 4:
                                    product3_values[4] = row_data[3]
                                self.shipping_table.item(product3_item, values=product3_values)
                                
                                product4_item = shipping_items[3]
                                product4_values = list(self.shipping_table.item(product4_item, "values"))
                                if len(row_data) >= 5:
                                    product4_values[1] = row_data[4]
                                if len(row_data) >= 6:
                                    product4_values[2] = row_data[5]
                                if len(row_data) >= 7:
                                    product4_values[3] = row_data[6]
                                if len(row_data) >= 8:
                                    product4_values[4] = row_data[7]
                                self.shipping_table.item(product4_item, values=product4_values)
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'产品运输变动费用（元）'
                start_match = re.search(start_pattern, page_content)
                
                if start_match:
                    start_pos = start_match.end()
                    
                    target_content = page_content[start_pos:]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_matches = re.findall(table_pattern, target_content, re.DOTALL)
                    
                    if len(table_matches) >= 1:
                        first_table_match = table_matches[0]
                        tbody_content = first_table_match
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        table_data = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            row_numbers = []
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                if clean_text:
                                    number_pattern = r'^[\d,]+\.?\d*$'
                                    if re.match(number_pattern, clean_text):
                                        clean_number = clean_text.replace(',', '')
                                        try:
                                            float_num = float(clean_number)
                                            row_numbers.append(clean_number)
                                        except ValueError:
                                            pass
                            
                            if row_numbers and tr_idx > 0:
                                table_data.append(row_numbers)
                        
                        if table_data:
                            shipping_items = self.shipping_table3.get_children()
                            
                            if len(table_data) >= 1 and len(shipping_items) >= 2:
                                row_data = table_data[0]
                                
                                product1_item = shipping_items[0]
                                product1_values = list(self.shipping_table3.item(product1_item, "values"))
                                if len(row_data) >= 1:
                                    product1_values[1] = row_data[0]
                                if len(row_data) >= 2:
                                    product1_values[2] = row_data[1]
                                if len(row_data) >= 3:
                                    product1_values[3] = row_data[2]
                                if len(row_data) >= 4:
                                    product1_values[4] = row_data[3]
                                self.shipping_table3.item(product1_item, values=product1_values)
                                
                                product2_item = shipping_items[1]
                                product2_values = list(self.shipping_table3.item(product2_item, "values"))
                                if len(row_data) >= 5:
                                    product2_values[1] = row_data[4]
                                if len(row_data) >= 6:
                                    product2_values[2] = row_data[5]
                                if len(row_data) >= 7:
                                    product2_values[3] = row_data[6]
                                if len(row_data) >= 8:
                                    product2_values[4] = row_data[7]
                                self.shipping_table3.item(product2_item, values=product2_values)
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'产品运输变动费用（元）'
                start_match = re.search(start_pattern, page_content)
                
                if start_match:
                    start_pos = start_match.end()
                    
                    target_content = page_content[start_pos:]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_matches = re.findall(table_pattern, target_content, re.DOTALL)
                    
                    if len(table_matches) >= 2:
                        second_table_match = table_matches[1]
                        tbody_content = second_table_match
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        table_data = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            row_numbers = []
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                if clean_text:
                                    number_pattern = r'^[\d,]+\.?\d*$'
                                    if re.match(number_pattern, clean_text):
                                        clean_number = clean_text.replace(',', '')
                                        try:
                                            float_num = float(clean_number)
                                            row_numbers.append(clean_number)
                                        except ValueError:
                                            pass
                            
                            if row_numbers and tr_idx > 0:
                                table_data.append(row_numbers)
                        
                        if table_data:
                            shipping_items = self.shipping_table3.get_children()
                            
                            if len(table_data) >= 1 and len(shipping_items) >= 4:
                                row_data = table_data[0]
                                
                                product3_item = shipping_items[2]
                                product3_values = list(self.shipping_table3.item(product3_item, "values"))
                                if len(row_data) >= 1:
                                    product3_values[1] = row_data[0]
                                if len(row_data) >= 2:
                                    product3_values[2] = row_data[1]
                                if len(row_data) >= 3:
                                    product3_values[3] = row_data[2]
                                if len(row_data) >= 4:
                                    product3_values[4] = row_data[3]
                                self.shipping_table3.item(product3_item, values=product3_values)
                                
                                product4_item = shipping_items[3]
                                product4_values = list(self.shipping_table3.item(product4_item, "values"))
                                if len(row_data) >= 5:
                                    product4_values[1] = row_data[4]
                                if len(row_data) >= 6:
                                    product4_values[2] = row_data[5]
                                if len(row_data) >= 7:
                                    product4_values[3] = row_data[6]
                                if len(row_data) >= 8:
                                    product4_values[4] = row_data[7]
                                self.shipping_table3.item(product4_item, values=product4_values)
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            try:
                page_content = page.content()
                import re
                
                
                start_pattern = r'<h4>原材料价格</h4>'
                start_match = re.search(start_pattern, page_content)
                
                if start_match:
                    start_pos = start_match.end()
                    
                    target_content = page_content[start_pos:]
                    
                    table_pattern = r'<table[^>]*>.*?<tbody>(.*?)</tbody>'
                    table_match = re.search(table_pattern, target_content, re.DOTALL)
                    
                    if table_match:
                        tbody_content = table_match.group(1)
                        
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        tr_matches = re.findall(tr_pattern, tbody_content, re.DOTALL)
                        
                        table_data = []
                        for tr_idx, tr_content in enumerate(tr_matches):
                            td_pattern = r'<td[^>]*>(.*?)</td>'
                            td_matches = re.findall(td_pattern, tr_content, re.DOTALL)
                            
                            row_numbers = []
                            for td_content in td_matches:
                                clean_text = re.sub(r'<[^>]+>', '', td_content)
                                clean_text = clean_text.strip()
                                
                                if clean_text:
                                    number_pattern = r'^[\d,]+\.?\d*$'
                                    if re.match(number_pattern, clean_text):
                                        clean_number = clean_text.replace(',', '')
                                        try:
                                            float_num = float(clean_number)
                                            row_numbers.append(clean_number)
                                        except ValueError:
                                            pass
                            
                            if row_numbers and tr_idx > 0:
                                table_data.append(row_numbers)
                        
                        if table_data:
                            discount_items = self.discount_table.get_children()
                            
                            for i, item in enumerate(discount_items):
                                current_values = list(self.discount_table.item(item, "values"))
                                
                                if i < len(table_data):
                                    row_data = table_data[i]
                                    if len(row_data) >= 1:
                                        current_values[0] = row_data[0]
                                    if len(row_data) >= 2:
                                        current_values[1] = row_data[1]
                                
                                self.discount_table.item(item, values=current_values)
                    else:
                        self.log(f"[调试] 目标区域前500字符: {target_content[:500]}")
            
            except Exception as e:
                import traceback
                self.log(f"[调试] 错误详情: {traceback.format_exc()}")
            
            return param_values
            
        except Exception as e:
            self.log(f"[错误] 提取规则参数失败: {e}")
            self.update_status("提取参数失败", color="red")
    
    def playwright_operation_loop(self):
        self.log("[Playwright] Playwright操作线程已启动")
        self.root.after(0, self.bring_to_front)
        while self.playwright_running:
            if self.playwright_queue:
                operation = self.playwright_queue.pop(0)
                op_type = operation[0]
                
                try:
                    if op_type == 'login':
                        _, username, password = operation
                        self.run_login_verification(username, password)
                    elif op_type == 'navigate':
                        _, game_url, game_id, game_name = operation
                        self.log(f"[Playwright] 执行导航: {game_id} - {game_name}")
                        
                        if self.page_handler.navigate(game_url):
                            if 'rules' in game_url:
                                self.root.after(0, lambda: self.update_status(f"已跳转到规则页面: {game_id} - {game_name}", color="green"))
                                self.playwright_queue.append(('extract_params',))
                            else:
                                self.root.after(0, lambda: self.update_status(f"已进入赛事: {game_id} - {game_name}", color="green"))
                                self.root.after(0, lambda: self.copy_button.config(state=tk.NORMAL))
                                self.root.after(0, lambda: self.import_button.config(state=tk.DISABLED))
                                self.root.after(0, lambda: self.extract_button.config(state=tk.DISABLED))
                                self.root.after(0, lambda: self.paste_button.config(state=tk.DISABLED))
                                self.root.after(0, self.clear_rules_table)
                                
                                try:
                                    page = self.page_handler.get_page()
                                    if page:
                                        import re
                                        page_content = page.content()
                                        pattern = r"您代表的公司是(.*?)。"
                                        match = re.search(pattern, page_content)
                                        if match:
                                            self.team_name = match.group(1).strip()
                                            self.log(f"[队伍] 检测到队伍名称为：{self.team_name}")
                                        else:
                                            self.log(f"[队伍] 未检测到队伍名称")
                                            self.team_name = ""
                                        
                                        teamid_pattern = r"function\s+teamid\s*\(\s*\)\s*\{\s*return\s+(\d+)\s*;\s*\}"
                                        teamid_match = re.search(teamid_pattern, page_content)
                                        if teamid_match:
                                            self.current_team_id = teamid_match.group(1)
                                            self.log(f"[队伍] 检测到team_id：{self.current_team_id}")
                                        else:
                                            self.log(f"[队伍] 未检测到team_id")
                                except Exception as e:
                                    self.log(f"[队伍] 提取队伍信息失败: {e}")
                                    self.team_name = ""
                                    
                            self.root.after(0, self.bring_to_front)
                            self.log(f"[成功] 成功跳转到页面")
                        else:
                            error_msg = "无法跳转到页面"
                            if not self.page_handler:
                                error_msg = "浏览器未启动，请先登录"
                            self.root.after(0, lambda: self.update_status("跳转失败", color="red"))
                            self.log("[错误] 无法跳转到页面")
                            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
                    elif op_type == 'extract_params':
                        param_values = self.extract_rules_parameters_in_thread()
                        self.root.after(0, lambda: self.update_status("规则提取完成，请打开Excel文件后点击导入规则", color="green"))
                    elif op_type == 'extract_periodid':
                        _, url = operation
                        self.log(f"[Playwright] 执行提取period_id: {url}")
                        
                        if not self.page_handler:
                            self.log("[错误] page_handler为None")
                            return
                        
                        page = self.page_handler.get_page()
                        if not page:
                            self.log("[错误] page为None")
                            return
                        
                        page.goto(url)
                        page.wait_for_load_state('networkidle')
                        
                        html_content = page.content()
                        
                        import re
                        pattern = r'<a[^>]*href="[^"]*periodid=(\d+)[^"]*"[^>]*>8</a>'
                        match = re.search(pattern, html_content)
                        
                        if match:
                            period_id = match.group(1)
                            report_url = f"https://www.ibizsim.cn/games/private_report?gameid={self.current_game_id}&periodid={period_id}&teamid={self.current_team_id}"
                            self.period_8_url = report_url
                            self.log(f"[Tab切换] 第8期报表完整URL: {report_url}")
                            self.root.after(0, lambda: self.extract_initial_report_button.config(state=tk.NORMAL))
                        else:
                            self.log("[Tab切换] 未找到第8期period_id")
                    elif op_type == 'navigate_period8':
                        _, url = operation
                        self.log(f"[Playwright] 执行跳转到第8期报表: {url}")
                        
                        if not self.page_handler:
                            self.log("[错误] page_handler为None")
                            return
                        
                        page = self.page_handler.get_page()
                        if not page:
                            self.log("[错误] page为None")
                            return
                        
                        page.goto(url)
                        page.wait_for_load_state('networkidle')
                        self.log(f"[报表] 已跳转到第8期报表页面")
                        self.root.after(0, lambda: self.update_status("已跳转到第8期报表页面", color="green"))
                    elif op_type == 'stop':
                        self.log("[Playwright] 执行停止操作")
                        self.cleanup_browser()
                        self.is_running = False
                        self.playwright_running = False
                        self.root.after(0, lambda: self.login_button.config(state=tk.NORMAL))
                        self.root.after(0, lambda: self.stop_button.config(state=tk.DISABLED))
                        self.root.after(0, self.clear_rules_table)
                        self.root.after(0, self.clear_inputs)
                    else:
                        self.log(f"[警告] 未知的操作类型: {op_type}")
                        
                except Exception as e:
                    self.root.after(0, lambda: self.log(f"[错误] Playwright操作失败: {e}"))
                    self.root.after(0, lambda: self.update_status("操作失败", color="red"))
                    self.root.after(0, lambda: messagebox.showerror("错误", f"Playwright操作失败：\n\n{e}"))
            else:
                import time
                time.sleep(0.1)
        
        self.root.after(0, lambda: self.log("[Playwright] Playwright操作线程已退出"))
    
    def start_verification(self):
        if self.is_running:
            messagebox.showwarning("警告", "验证正在进行中，请先停止当前任务")
            return
        
        username = self.username.get().strip()
        password = self.password.get().strip()
        
        if not username or not password:
            messagebox.showerror("错误", "用户名和密码不能为空")
            return
        
        self.is_running = True
        self.login_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        
        self.update_status("正在启动验证...", color="blue")
        
        if not self.playwright_thread or not self.playwright_thread.is_alive():
            self.playwright_running = True
            self.playwright_thread = threading.Thread(target=self.playwright_operation_loop)
            self.playwright_thread.daemon = True
            self.playwright_thread.start()
        
        self.playwright_queue.append(('login', username, password))
        self.log(f"[队列] 已添加登录请求到Playwright队列")
    
    def stop_verification(self):
        if not self.is_running:
            return
        
        self.update_status("正在停止...", color="orange")
        self.log("[操作] 用户请求停止验证")
        
        self.playwright_queue.append(('stop',))
    
    def run_login_verification(self, username, password):
        try:
            self.log("[配置] 加载配置文件...")
            self.settings = Settings()
            self.log(f"[配置] 网站: {self.settings.website.get('base_url')}")
            self.log(f"[配置] 登录页面: {self.settings.website.get('login_url')}")
            
            self.settings.username = username
            self.settings.password = password
            
            self.log("[浏览器] 启动浏览器...")
            browser_config = self.settings.browser
            headless = browser_config.get('headless', False)
            timeout = browser_config.get('timeout', 30000)
            screenshot_on_error = browser_config.get('screenshot_on_error', True)
            
            self.browser_manager = BrowserManager(
                headless=headless,
                timeout=timeout,
                screenshot_on_error=screenshot_on_error
            )
            
            if not self.browser_manager.start():
                self.update_status("浏览器启动失败", color="red")
                self.log("[错误] 无法启动浏览器")
                self.is_running = False
                self.root.after(0, lambda: self.login_button.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.stop_button.config(state=tk.DISABLED))
                return
            
            self.log("[浏览器] 浏览器启动成功")
            self.root.after(0, self.bring_to_front)
            
            page = self.browser_manager.get_page()
            if not page:
                self.update_status("无法获取页面对象", color="red")
                self.log("[错误] 无法获取页面对象")
                self.cleanup_browser()
                self.is_running = False
                self.root.after(0, lambda: self.login_button.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.stop_button.config(state=tk.DISABLED))
                return
            
            self.page_handler = PageHandler(page, on_navigate=self.bring_to_front)
            self.settings.username = username
            self.settings.password = password
            
            self.login_handler = LoginHandler(self.page_handler, self.settings, on_navigate=self.bring_to_front)
            self.log("[登录] 初始化登录处理器")
            
            self.update_status("正在登录...", color="blue")
            self.log("[登录] 导航到登录页面")
            self.log("[登录] 填写用户名和密码")
            
            self.log(f"[调试] 开始调用登录处理器...")
            login_success = self.login_handler.login()
            self.log(f"[调试] 登录处理器返回值: {login_success}")
            
            if login_success:
                self.update_status("登录成功！", color="green")
                self.log("[成功] 登录验证成功！")
                self.log(f"[状态] 已登录: {self.login_handler.is_authenticated()}")
                self.log(f"[页面] 当前页面: {page.url}")
                
                self.root.after(0, self.bring_to_front)
                
                self.update_status("正在导航到赛事列表页面...", color="blue")
                
                mygames_url = "https://www.ibizsim.cn/games/mygames"
                if self.page_handler.navigate(mygames_url):
                    self.root.after(0, self.bring_to_front)
                    self.update_status("正在加载赛事列表...", color="blue")
                    
                    self.load_games()
                    self.root.after(0, self.bring_to_front)
                else:
                    self.log("[错误] 无法导航到赛事列表页面")
                    self.update_status("导航到赛事列表失败", color="red")
            else:
                self.log("[调试] 进入登录失败分支")
                self.root.after(0, lambda: self.update_status("登录失败！请检查用户名和密码", color="red"))
                self.log("[失败] 登录验证失败")
                self.log("[原因] 可能是用户名或密码错误")
                
                self.root.after(0, lambda: messagebox.showerror(
                    "登录失败",
                    "登录验证失败！\n\n可能的原因：\n1. 用户名或密码错误\n2. 网站选择器配置错误\n3. 网络连接问题\n4. 验证码输入错误\n\n请检查输入信息后重试，或点击\"查看日志\"按钮查看详细错误信息。"
                ))
                
                self.is_running = False
                self.root.after(0, lambda: self.login_button.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.stop_button.config(state=tk.DISABLED))
                
        except Exception as e:
            self.update_status(f"发生错误: {str(e)}", color="red")
            self.log(f"[异常] 测试过程中发生异常: {e}")
            messagebox.showerror("错误", f"发生错误：\n\n{e}")
            self.is_running = False
        finally:
            if not self.is_running:
                self.cleanup_browser()
                self.root.after(0, lambda: self.login_button.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.stop_button.config(state=tk.DISABLED))
    
    def cleanup_browser(self):
        self.playwright_running = False
        if self.browser_manager:
            self.log("[清理] 关闭浏览器...")
            try:
                self.browser_manager.stop()
            except Exception:
                pass
            finally:
                self.browser_manager = None
                if not self.page_handler or not self.is_running:
                    self.page_handler = None
                self.login_handler = None
    
    def on_closing(self):
        if messagebox.askyesno("确认退出", "确定要退出 iBizSim 助手吗？"):
            self.cleanup_browser()
            if self.is_running:
                self.stop_verification()
            self.root.destroy()
    
    def run(self):
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()


def main():
    app = LoginGUI()
    app.run()


if __name__ == '__main__':
    main()