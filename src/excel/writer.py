from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from typing import Any, List, Dict, Optional
from pathlib import Path
from src.utils.logger import get_logger


class ExcelWriter:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.logger = get_logger()
    
    def create_new(self):
        try:
            self.workbook = Workbook()
            self.logger.info(f"Created new workbook: {self.file_path}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to create new workbook: {e}")
            return False
    
    def open_existing(self):
        try:
            if not self.file_path.exists():
                self.logger.warning(f"File does not exist, creating new: {self.file_path}")
                return self.create_new()
            
            self.workbook = load_workbook(self.file_path)
            self.logger.info(f"Opened existing workbook: {self.file_path}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to open workbook: {e}")
            return False
    
    def save(self):
        try:
            if not self.workbook:
                self.logger.error("No workbook to save")
                return False
            
            self.file_path.parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(str(self.file_path))
            self.logger.info(f"Workbook saved: {self.file_path}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to save workbook: {e}")
            return False
    
    def close(self):
        try:
            if self.workbook:
                self.workbook.close()
                self.workbook = None
                self.logger.info("Workbook closed")
        except Exception as e:
            self.logger.error(f"Error closing workbook: {e}")
    
    def get_sheet(self, sheet_name: str):
        if not self.workbook:
            self.logger.error("Workbook not opened")
            return None
        
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            return self.create_sheet(sheet_name)
    
    def create_sheet(self, sheet_name: str):
        try:
            if not self.workbook:
                self.logger.error("Workbook not opened")
                return None
            
            sheet = self.workbook.create_sheet(title=sheet_name)
            self.logger.info(f"Created new sheet: {sheet_name}")
            return sheet
        except Exception as e:
            self.logger.error(f"Failed to create sheet: {e}")
            return None
    
    def write_cell(self, cell_reference: str, value: Any, 
                   sheet_name: str = 'Sheet1'):
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return False
            
            sheet[cell_reference] = value
            self.logger.debug(f"Written {value} to {cell_reference}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to write to cell {cell_reference}: {e}")
            return False
    
    def write_row(self, row_number: int, data: List[Any], 
                  sheet_name: str = 'Sheet1', start_column: int = 1):
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return False
            
            for i, value in enumerate(data):
                sheet.cell(row=row_number, column=start_column + i, value=value)
            
            self.logger.debug(f"Written row {row_number}: {data}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to write row {row_number}: {e}")
            return False
    
    def write_column(self, column_letter: str, data: List[Any], 
                     sheet_name: str = 'Sheet1', start_row: int = 1):
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return False
            
            for i, value in enumerate(data):
                sheet[f"{column_letter}{start_row + i}"] = value
            
            self.logger.debug(f"Written column {column_letter}: {len(data)} cells")
            return True
        except Exception as e:
            self.logger.error(f"Failed to write column {column_letter}: {e}")
            return False
    
    def write_range(self, start_cell: str, data: List[List[Any]], 
                   sheet_name: str = 'Sheet1'):
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return False
            
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            start_col = int(''.join(filter(str.isalpha, start_cell)), 36) - 9
            
            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    sheet.cell(row=start_row + i, column=start_col + j, value=value)
            
            self.logger.debug(f"Written range starting at {start_cell}: {len(data)} rows")
            return True
        except Exception as e:
            self.logger.error(f"Failed to write range {start_cell}: {e}")
            return False
    
    def append_row(self, data: List[Any], sheet_name: str = 'Sheet1'):
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return False
            
            sheet.append(data)
            self.logger.debug(f"Appended row: {data}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to append row: {e}")
            return False
    
    def write_dict(self, data: Dict[str, Any], sheet_name: str = 'Sheet1',
                   start_row: int = 1, start_column: int = 1):
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return False
            
            for i, (key, value) in enumerate(data.items()):
                sheet.cell(row=start_row, column=start_column + i, value=key)
                sheet.cell(row=start_row + 1, column=start_column + i, value=value)
            
            self.logger.debug(f"Written dictionary: {len(data)} fields")
            return True
        except Exception as e:
            self.logger.error(f"Failed to write dictionary: {e}")
            return False
    
    def write_header(self, headers: List[str], sheet_name: str = 'Sheet1',
                     row: int = 1, bold: bool = True):
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return False
            
            for i, header in enumerate(headers):
                cell = sheet.cell(row=row, column=i + 1, value=header)
                if bold:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
            
            self.logger.debug(f"Written headers: {headers}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to write headers: {e}")
            return False
    
    def clear_sheet(self, sheet_name: str):
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return False
            
            sheet.delete_rows(1, sheet.max_row)
            self.logger.info(f"Cleared sheet: {sheet_name}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to clear sheet: {e}")
            return False
    
    def get_last_row(self, sheet_name: str = 'Sheet1') -> int:
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return 0
            
            return sheet.max_row
        except Exception as e:
            self.logger.error(f"Failed to get last row: {e}")
            return 0
    
    def __enter__(self):
        self.open_existing()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()
        self.close()
