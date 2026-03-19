from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Any, List, Dict, Optional
from pathlib import Path
from src.utils.logger import get_logger


class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.logger = get_logger()
    
    def open(self):
        try:
            if not self.file_path.exists():
                raise FileNotFoundError(f"Excel file not found: {self.file_path}")
            
            self.workbook = load_workbook(self.file_path)
            self.logger.info(f"Excel file opened: {self.file_path}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to open Excel file: {e}")
            return False
    
    def close(self):
        try:
            if self.workbook:
                self.workbook.close()
                self.workbook = None
                self.logger.info("Excel file closed")
        except Exception as e:
            self.logger.error(f"Error closing Excel file: {e}")
    
    def get_sheet_names(self) -> List[str]:
        if not self.workbook:
            self.logger.error("Workbook not opened")
            return []
        return self.workbook.sheetnames
    
    def get_sheet(self, sheet_name: str):
        if not self.workbook:
            self.logger.error("Workbook not opened")
            return None
        
        if sheet_name not in self.workbook.sheetnames:
            self.logger.error(f"Sheet not found: {sheet_name}")
            return None
        
        return self.workbook[sheet_name]
    
    def read_cell(self, cell_reference: str, sheet_name: str = 'Sheet1') -> Optional[Any]:
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return None
            
            cell = sheet[cell_reference]
            value = cell.value
            
            self.logger.debug(f"Read cell {cell_reference}: {value}")
            return value
            
        except Exception as e:
            self.logger.error(f"Failed to read cell {cell_reference}: {e}")
            return None
    
    def read_row(self, row_number: int, sheet_name: str = 'Sheet1') -> List[Any]:
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return []
            
            row_data = []
            for cell in sheet[row_number]:
                row_data.append(cell.value)
            
            self.logger.debug(f"Read row {row_number}: {row_data}")
            return row_data
            
        except Exception as e:
            self.logger.error(f"Failed to read row {row_number}: {e}")
            return []
    
    def read_column(self, column_letter: str, sheet_name: str = 'Sheet1') -> List[Any]:
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return []
            
            column_data = []
            for cell in sheet[column_letter]:
                column_data.append(cell.value)
            
            self.logger.debug(f"Read column {column_letter}: {len(column_data)} cells")
            return column_data
            
        except Exception as e:
            self.logger.error(f"Failed to read column {column_letter}: {e}")
            return []
    
    def read_range(self, start_cell: str, end_cell: str, 
                   sheet_name: str = 'Sheet1') -> List[List[Any]]:
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return []
            
            range_data = []
            for row in sheet[start_cell:end_cell]:
                row_data = [cell.value for cell in row]
                range_data.append(row_data)
            
            self.logger.debug(f"Read range {start_cell}:{end_cell}: {len(range_data)} rows")
            return range_data
            
        except Exception as e:
            self.logger.error(f"Failed to read range {start_cell}:{end_cell}: {e}")
            return []
    
    def read_all_data(self, sheet_name: str = 'Sheet1') -> List[List[Any]]:
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return []
            
            all_data = []
            for row in sheet.iter_rows(values_only=True):
                all_data.append(list(row))
            
            self.logger.info(f"Read all data from sheet {sheet_name}: {len(all_data)} rows")
            return all_data
            
        except Exception as e:
            self.logger.error(f"Failed to read all data: {e}")
            return []
    
    def read_to_dict(self, sheet_name: str = 'Sheet1', 
                    header_row: int = 1) -> List[Dict[str, Any]]:
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return []
            
            headers = [cell.value for cell in sheet[header_row]]
            data = []
            
            for row in sheet.iter_rows(min_row=header_row + 1):
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = cell.value
                data.append(row_dict)
            
            self.logger.info(f"Read {len(data)} rows as dictionaries")
            return data
            
        except Exception as e:
            self.logger.error(f"Failed to read to dictionary: {e}")
            return []
    
    def get_cell_value(self, row: int, column: int, 
                      sheet_name: str = 'Sheet1') -> Optional[Any]:
        try:
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                return None
            
            cell = sheet.cell(row=row, column=column)
            value = cell.value
            
            self.logger.debug(f"Read cell ({row}, {column}): {value}")
            return value
            
        except Exception as e:
            self.logger.error(f"Failed to read cell ({row}, {column}): {e}")
            return None
    
    def __enter__(self):
        self.open()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
